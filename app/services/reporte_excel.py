import os
import uuid
import urllib.request
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session

from app.core.config import settings
from app.repositories.equipo import EquipoRepository
from app.repositories.area import AreaRepository
from app.repositories.clasificacion import ClasificacionRepository
from app.repositories.foto import FotoRepository
from app.repositories.mantenimiento import MantenimientoRepository
from app.repositories.usuario import UsuarioRepository
from app.models.enums import TipoFotoEnum

LOGO_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "assets", "logo.jpg",
)


def _open_image_bytes(url_or_path: str) -> BytesIO | None:
    if url_or_path.startswith("http"):
        try:
            with urllib.request.urlopen(url_or_path, timeout=10) as resp:
                return BytesIO(resp.read())
        except Exception:
            return None
    local = url_or_path.lstrip("/")
    if os.path.exists(local):
        with open(local, "rb") as f:
            return BytesIO(f.read())
    return None


# ── ESTILOS ─────────────────────────────────────────

VERDE_OSCURO = PatternFill(
    start_color="1F4E79",
    end_color="1F4E79",
    fill_type="solid",
)
VERDE_CLARO = PatternFill(
    start_color="2E75B6",
    end_color="2E75B6",
    fill_type="solid",
)
GRIS_CLARO = PatternFill(
    start_color="F5F5F5",
    end_color="F5F5F5",
    fill_type="solid",
)

FUENTE_HEADER = Font(
    name="Calibri",
    size=11,
    bold=True,
    color="FFFFFF",
)
FUENTE_NORMAL = Font(
    name="Calibri",
    size=10,
)
FUENTE_TITULO = Font(
    name="Calibri",
    size=14,
    bold=True,
    color="1F4E79",
)

BORDE = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ALINEAR_CENTRO = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)


def _aplicar_estilo_header(ws, fila, num_cols):
    for col in range(1, num_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = FUENTE_HEADER
        celda.fill = VERDE_OSCURO
        celda.alignment = ALINEAR_CENTRO
        celda.border = BORDE


def _aplicar_estilo_fila(ws, fila, num_cols, alternar=False):
    for col in range(1, num_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = FUENTE_NORMAL
        celda.border = BORDE
        celda.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
        if alternar:
            celda.fill = GRIS_CLARO


def _ajustar_anchos(ws, anchos):
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


# ── REPORTE DE EQUIPOS EXCEL ────────────────────────

def generar_excel_equipos(
    session: Session,
    area_id: uuid.UUID | None = None,
    clasificacion_id: uuid.UUID | None = None,
    estado: str | None = None,
    nivel: int | None = None,
) -> BytesIO:
    """Genera Excel con listado de equipos y filtros opcionales."""

    equipo_repo = EquipoRepository(session)
    area_repo = AreaRepository(session)
    clasif_repo = ClasificacionRepository(session)

    # Obtener equipos con filtros
    equipos = equipo_repo.filter(
        area_id=area_id,
        clasificacion_id=clasificacion_id,
        estado=estado,
        nivel=nivel,
        limit=1000,
    )

    wb = Workbook()

    # ── HOJA 1: LISTADO DE EQUIPOS ──────────────────
    ws = wb.active
    ws.title = "Equipos"

    # ── Logo ──
    row = 1
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 1606
        img.height = 57
        img.anchor = f"A{row}"
        ws.add_image(img)
        row += 3

    # Título
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1, value="LISTADO DE EQUIPOS - SISTEMA DE FICHAS TÉCNICAS")
    ws.cell(row=row, column=1).font = FUENTE_TITULO
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1, value=(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Total: {len(equipos)} equipos"
    ))
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=9, color="999999")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Headers
    headers = [
        "Código", "Nombre", "Nivel", "Estado",
        "Clasificación", "Área", "Marca", "Modelo",
        "N° Serie", "Potencia", "Padre", "Fecha Creación",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _aplicar_estilo_header(ws, row, len(headers))
    row += 1

    # Datos
    for i, eq in enumerate(equipos):
        fila = i + row

        nombre_area = ""
        if eq.area_id:
            area = area_repo.get_by_id(eq.area_id)
            nombre_area = area.nombre if area else ""

        nombre_clasif = ""
        if eq.clasificacion_id:
            clasif = clasif_repo.get_by_id(eq.clasificacion_id)
            nombre_clasif = clasif.nombre if clasif else ""

        nombre_padre = ""
        if eq.equipo_padre_id:
            padre = equipo_repo.get_by_id(eq.equipo_padre_id)
            nombre_padre = padre.codigo_equipo if padre else ""

        datos = [
            eq.codigo_equipo,
            eq.nombre,
            eq.nivel,
            eq.estado.value,
            nombre_clasif,
            nombre_area,
            eq.marca or "",
            eq.modelo or "",
            eq.numero_serie or "",
            eq.potencia or "",
            nombre_padre,
            eq.created_at.strftime("%d/%m/%Y") if eq.created_at else "",
        ]

        for col, valor in enumerate(datos, 1):
            ws.cell(row=fila, column=col, value=valor)

        _aplicar_estilo_fila(
            ws, fila, len(headers),
            alternar=(i % 2 == 0),
        )

    _ajustar_anchos(ws, [
        15, 30, 6, 12, 18, 20, 15, 15, 18, 12, 18, 14,
    ])

    # ── HOJA 2: RESUMEN POR ESTADO ──────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "RESUMEN POR ESTADO"
    ws2["A1"].font = FUENTE_TITULO

    resumen_headers = ["Estado", "Cantidad", "Porcentaje"]
    for col, h in enumerate(resumen_headers, 1):
        ws2.cell(row=3, column=col, value=h)
    _aplicar_estilo_header(ws2, 3, 3)

    estados_contados = {}
    for eq in equipos:
        est = eq.estado.value
        estados_contados[est] = estados_contados.get(est, 0) + 1

    total = len(equipos) or 1
    for i, (estado, cant) in enumerate(estados_contados.items()):
        fila = i + 4
        ws2.cell(row=fila, column=1, value=estado)
        ws2.cell(row=fila, column=2, value=cant)
        ws2.cell(
            row=fila, column=3,
            value=f"{(cant / total) * 100:.1f}%",
        )
        _aplicar_estilo_fila(ws2, fila, 3, alternar=(i % 2 == 0))

    _ajustar_anchos(ws2, [20, 12, 12])

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── HELPER: Construir ficha estructurada ──────────────────

def _build_ficha_estructurada(equipo, nombre_area, nombre_clasif, nombre_padre, hijos):
    """Convierte un equipo ORM en el formato estructurado de ficha técnica."""
    secciones = []

    # ── Ubicaciones del Equipo ──
    ubi = {"nombre": "Ubicaciones del Equipo", "campos": []}
    ubi["campos"].append({"clave": "Código", "valor": equipo.codigo_equipo})
    if nombre_clasif:
        ubi["campos"].append({"clave": "Clasificación", "valor": nombre_clasif})
    if nombre_area:
        ubi["campos"].append({"clave": "Área / Ubicación", "valor": nombre_area})
    ubi["campos"].append({"clave": "Nivel Jerárquico", "valor": str(equipo.nivel)})
    if nombre_padre:
        ubi["campos"].append({"clave": "Equipo Padre", "valor": nombre_padre})
    secciones.append(ubi)

    # ── Datos del Fabricante ──
    fab = {"nombre": "Datos del Fabricante", "campos": []}
    for k, v in [
        ("Marca", equipo.marca),
        ("Modelo", equipo.modelo),
        ("N° Serie", equipo.numero_serie),
        ("Año Fabricación", str(equipo.anio_fabricacion) if equipo.anio_fabricacion else None),
        ("Proveedor", equipo.proveedor),
    ]:
        if v:
            fab["campos"].append({"clave": k, "valor": v})
    if fab["campos"]:
        secciones.append(fab)

    # ── Datos de Placa / Datasheet ──
    pla = {"nombre": "Datos de Placa / Datasheet", "campos": []}
    for k, v in [
        ("Potencia", equipo.potencia),
        ("Voltaje", equipo.voltaje),
        ("RPM", equipo.rpm),
        ("Capacidad", equipo.capacidad),
    ]:
        if v:
            pla["campos"].append({"clave": k, "valor": v})
    if pla["campos"]:
        secciones.append(pla)

    # ── Datos de la Unidad ──
    uni = {"nombre": "Datos de la Unidad", "campos": []}
    if equipo.fecha_adquisicion:
        uni["campos"].append({
            "clave": "Fecha de Adquisición",
            "valor": equipo.fecha_adquisicion.strftime("%d/%m/%Y"),
        })
    uni["campos"].append({"clave": "Estado", "valor": equipo.estado.value})
    if equipo.motivo_estado:
        uni["campos"].append({"clave": "Motivo de Estado", "valor": equipo.motivo_estado})
    if equipo.created_at:
        uni["campos"].append({
            "clave": "Fecha de Registro",
            "valor": equipo.created_at.strftime("%d/%m/%Y"),
        })
    if equipo.descripcion:
        uni["campos"].append({"clave": "Descripción", "valor": equipo.descripcion})
    secciones.append(uni)

    # ── Sub-Equipos ──
    if hijos:
        subs = {"nombre": "Sub-Equipos / Componentes", "campos": []}
        for s in hijos:
            subs["campos"].append({
                "clave": s.codigo_equipo,
                "valor": f"{s.nombre} — {s.estado.value}",
            })
        secciones.append(subs)

    # ── Observaciones ──
    if equipo.observaciones:
        secciones.append({
            "nombre": "Observaciones",
            "campos": [{"clave": "Observaciones", "valor": equipo.observaciones}],
        })

    return {
        "empresa": settings.APP_NAME,
        "titulo": "Ficha Técnica del Equipo",
        "funcion": "",
        "equipo": equipo.nombre,
        "tag": equipo.codigo_equipo,
        "secciones": secciones,
    }


# ── FICHA TÉCNICA INDIVIDUAL EXCEL ─────────────────

def generar_ficha_equipo_excel(
    session: Session,
    equipo_id: uuid.UUID,
) -> BytesIO:
    """Genera Excel con la ficha técnica detallada de un equipo."""

    equipo_repo = EquipoRepository(session)
    area_repo = AreaRepository(session)
    clasif_repo = ClasificacionRepository(session)

    equipo = equipo_repo.get_by_id(equipo_id)
    if not equipo:
        raise ValueError("Equipo no encontrado")

    # Resolver nombres
    nombre_area = area_repo.get_by_id(equipo.area_id).nombre if equipo.area_id else ""
    nombre_clasif = clasif_repo.get_by_id(equipo.clasificacion_id).nombre if equipo.clasificacion_id else ""
    nombre_padre = ""
    if equipo.equipo_padre_id:
        padre = equipo_repo.get_by_id(equipo.equipo_padre_id)
        nombre_padre = padre.nombre if padre else ""
    hijos = equipo_repo.get_hijos_directos(equipo.id)

    ficha = _build_ficha_estructurada(
        equipo, nombre_area, nombre_clasif, nombre_padre, hijos,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Técnica"

    # Estilos
    azul_oscuro = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    azul_medio = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    gris_alterno = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    blanco = PatternFill()

    hdr_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    tit_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    sec_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    key_font = Font(name="Calibri", size=10, bold=True)
    val_font = Font(name="Calibri", size=10)

    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    centro = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50

    row = 1

    # ── Logo ──
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 689
        img.height = 19
        img.anchor = f"A{row}"
        ws.add_image(img)
        row += 3

    # ── Encabezado ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value=ficha["empresa"])
    c.font = hdr_font
    c.fill = azul_oscuro
    c.alignment = centro
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value=ficha["titulo"])
    c.font = tit_font
    c.alignment = centro
    row += 1

    # Equipo y TAG
    ws.cell(row=row, column=1, value="Equipo").font = key_font
    ws.cell(row=row, column=1).border = borde
    ws.cell(row=row, column=2, value=ficha["equipo"]).font = val_font
    ws.cell(row=row, column=2).border = borde
    row += 1

    ws.cell(row=row, column=1, value="TAG").font = key_font
    ws.cell(row=row, column=1).border = borde
    ws.cell(row=row, column=2, value=ficha["tag"]).font = Font(
        name="Calibri", size=10, bold=True, color="C00000",
    )
    ws.cell(row=row, column=2).border = borde
    row += 1

    # ── Secciones ──
    for seccion in ficha["secciones"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=1, value=seccion["nombre"])
        c.font = sec_font
        c.fill = azul_medio
        c.alignment = centro
        row += 1

        for i, campo in enumerate(seccion["campos"]):
            fill = gris_alterno if i % 2 == 0 else blanco
            c1 = ws.cell(row=row, column=1, value=campo["clave"])
            c1.font = key_font
            c1.border = borde
            c1.fill = fill

            c2 = ws.cell(row=row, column=2, value=campo["valor"] if campo["valor"] else "—")
            c2.font = val_font
            c2.border = borde
            c2.fill = fill
            row += 1

        row += 1

    # ── Foto del Equipo ──
    foto_repo = FotoRepository(session)
    fotos = foto_repo.get_by_equipo_y_tipo(equipo_id, TipoFotoEnum.GENERAL)
    if fotos:
        foto_bytes = _open_image_bytes(fotos[0].url)
        if foto_bytes:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row=row, column=1, value="Foto del Equipo")
            c.font = sec_font
            c.fill = azul_medio
            c.alignment = centro
            row += 1
            foto_img = XLImage(foto_bytes)
            foto_img.width = 400
            foto_img.height = 300
            foto_img.anchor = f"A{row}"
            ws.add_image(foto_img)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── REPORTE DE MANTENIMIENTOS EXCEL ─────────────────

def generar_excel_mantenimientos(
    session: Session,
    equipo_id: uuid.UUID | None = None,
    fecha_desde=None,
    fecha_hasta=None,
) -> BytesIO:
    """Genera Excel con listado de mantenimientos."""

    mantto_repo = MantenimientoRepository(session)
    equipo_repo = EquipoRepository(session)
    usuario_repo = UsuarioRepository(session)

    if equipo_id:
        mantenimientos = mantto_repo.get_by_equipo(
            equipo_id, limit=1000,
        )
    else:
        mantenimientos = mantto_repo.filter(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            limit=1000,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"

    # ── Logo ──
    row = 1
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 1440
        img.height = 57
        img.anchor = f"A{row}"
        ws.add_image(img)
        row += 3

    # Título
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="HISTORIAL DE MANTENIMIENTOS")
    ws.cell(row=row, column=1).font = FUENTE_TITULO
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Total: {len(mantenimientos)} registros"
    ))
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=9, color="999999")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Headers
    headers = [
        "Fecha", "Tipo", "Título", "Estado",
        "Equipo", "Código Equipo", "Realizado por",
        "Descripción",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _aplicar_estilo_header(ws, row, len(headers))
    row += 1

    # Datos
    for i, m in enumerate(mantenimientos):
        fila = i + row

        equipo = equipo_repo.get_by_id(m.equipo_id)
        nombre_equipo = equipo.nombre if equipo else ""
        codigo_equipo = equipo.codigo_equipo if equipo else ""

        nombre_usuario = m.realizado_por if m.realizado_por else ""
        if not nombre_usuario and m.usuario_id:
            usuario = usuario_repo.get_by_id(m.usuario_id)
            nombre_usuario = usuario.nombre if usuario else ""

        datos = [
            str(m.fecha),
            m.tipo.value,
            m.titulo,
            m.estado.value,
            nombre_equipo,
            codigo_equipo,
            nombre_usuario,
            (m.descripcion or "")[:100],
        ]

        for col, valor in enumerate(datos, 1):
            ws.cell(row=fila, column=col, value=valor)

        _aplicar_estilo_fila(
            ws, fila, len(headers),
            alternar=(i % 2 == 0),
        )

    _ajustar_anchos(ws, [12, 12, 35, 12, 25, 18, 18, 40])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── REPORTE DE REPUESTOS EXCEL ──────────────────────

def generar_excel_repuestos(
    session: Session,
) -> BytesIO:
    """Genera Excel con todos los repuestos (nivel 4+)."""

    equipo_repo = EquipoRepository(session)

    # Buscar todos los equipos que son repuestos (nivel >= 4)
    todos = equipo_repo.get_all(limit=5000)
    repuestos = [e for e in todos if e.nivel >= 4]

    wb = Workbook()
    ws = wb.active
    ws.title = "Repuestos"

    # ── Logo ──
    row = 1
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 1398
        img.height = 61
        img.anchor = f"A{row}"
        ws.add_image(img)
        row += 3

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="CATÁLOGO DE REPUESTOS")
    ws.cell(row=row, column=1).font = FUENTE_TITULO
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value=(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Total: {len(repuestos)} repuestos"
    ))
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=9, color="999999")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    headers = [
        "Código", "Nombre", "Descripción",
        "Componente Padre", "Equipo Raíz", "Estado",
        "Observaciones",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _aplicar_estilo_header(ws, row, len(headers))
    row += 1

    for i, r in enumerate(repuestos):
        fila = i + row

        # Resolver padre
        nombre_padre = ""
        if r.equipo_padre_id:
            padre = equipo_repo.get_by_id(r.equipo_padre_id)
            nombre_padre = padre.nombre if padre else ""

        # Resolver equipo raíz (subir hasta nivel 1)
        nombre_raiz = ""
        actual = r
        while actual.equipo_padre_id:
            padre = equipo_repo.get_by_id(actual.equipo_padre_id)
            if padre:
                actual = padre
            else:
                break
        if actual.nivel == 1:
            nombre_raiz = actual.nombre

        datos = [
            r.codigo_equipo,
            r.nombre,
            r.descripcion or "",
            nombre_padre,
            nombre_raiz,
            r.estado.value,
            r.observaciones or "",
        ]

        for col, valor in enumerate(datos, 1):
            ws.cell(row=fila, column=col, value=valor)

        _aplicar_estilo_fila(
            ws, fila, len(headers),
            alternar=(i % 2 == 0),
        )

    _ajustar_anchos(ws, [15, 30, 30, 25, 25, 12, 30])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
